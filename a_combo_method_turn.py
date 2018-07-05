

from evaluator import Evaluator
from deck import Deck
from openpyxl import load_workbook
from card import Card

# 1. Initiate class
evaluator = Evaluator()

# 2. Workbooks used for output and input

# 2a Input workbook is range_lookup.xlsm
wb_input = load_workbook(filename='combo_method.xlsm', data_only=True)
input_sheet = wb_input['turn']

# 2b Output workbook
dest_filename = 'combo_method_output.xlsx'
wb_output = load_workbook(filename=dest_filename, read_only=False)
ws1 = wb_output['turn']
ws_all_combos = wb_input['all_combos']


output_counter = ws1.max_row + 1

set_deuces_rating = 2467  # ranking of threedeuces

#  3. Load in hero and flop cards

combo_count = input_sheet.cell(row=2, column=16).value

for l in range(5, combo_count + 5):

    # get flop cards
    flop_lookup = input_sheet.cell(row=l, column=19).value
    flop_card_1 = input_sheet.cell(row=l, column=20).value
    flop_card_2 = input_sheet.cell(row=l, column=21).value
    flop_card_3 = input_sheet.cell(row=l, column=22).value
    flop_cards = [flop_card_1, flop_card_2, flop_card_3]

    hero_card_group = input_sheet.cell(row=l, column=15).value
    hero_cards = input_sheet.cell(row=l, column=16).value
    hero_card_1 = input_sheet.cell(row=l, column=17).value
    hero_card_2 = input_sheet.cell(row=l, column=18).value
    hero_hand = [hero_card_1, hero_card_2]
    hero_ranks = [Card.get_rank_int(hero_card_1), Card.get_rank_int(hero_card_2)]

    turn_card_str = input_sheet.cell(row=l, column=23).value
    turn_card = [input_sheet.cell(row=l, column=24).value]

    flop_and_hero_and_turn_cards = [hero_card_1, hero_card_2, flop_card_1, flop_card_2, flop_card_3, turn_card[0]]

    for j in range(2, 1327):
        villain_combo = ws_all_combos.cell(row=j, column=2).value
        villain_actual = ws_all_combos.cell(row=j, column=3).value
        villain_card_1 = ws_all_combos.cell(row=j, column=6).value
        villain_card_2 = ws_all_combos.cell(row=j, column=7).value
        villain_hand = [villain_card_1, villain_card_2]

        if villain_card_1 not in flop_and_hero_and_turn_cards and villain_card_2 not in flop_and_hero_and_turn_cards:

            turn_board = flop_cards + turn_card

            hero_best_turn = 0
            hero_best_river = 0
            turn_tie = 0
            river_tie = 0

            flop_flush_draw_count = 0
            flop_straight_draw_count = 0

            ## end new

            ###### drawing power  ###############

            # turn
            hero_turn_air = 0
            hero_turn_pair = 0
            hero_turn_top_pair = 0
            hero_turn_overpair = 0
            hero_turn_two_pair = 0
            hero_turn_trips = 0
            hero_turn_straight = 0
            hero_turn_flush = 0
            hero_turn_full_house = 0
            hero_turn_four_of_a_kind = 0
            hero_turn_straight_flush = 0

            villain_turn_air = 0
            villain_turn_full_house = 0
            villain_turn_top_pair = 0
            villain_turn_overpair = 0
            villain_turn_pair = 0
            villain_turn_two_pair = 0
            villain_turn_trips = 0
            villain_turn_straight = 0
            villain_turn_flush = 0
            villain_turn_four_of_a_kind = 0
            villain_turn_straight_flush = 0

            # river
            hero_river_air = 0
            hero_river_pair = 0
            hero_river_top_pair = 0
            hero_river_overpair = 0
            hero_river_two_pair = 0
            hero_river_trips = 0
            hero_river_straight = 0
            hero_river_flush = 0
            hero_river_full_house = 0
            hero_river_four_of_a_kind = 0
            hero_river_straight_flush = 0

            # hand strength

            hero_villain_count_pair = 0
            hero_villain_count_top_pair = 0
            hero_villain_count_two_pair = 0
            hero_villain_count_overpair = 0
            hero_villain_count_trips = 0
            hero_villain_count_straight = 0
            hero_villain_count_flush = 0
            hero_villain_count_full_house = 0
            hero_villain_count_four_of_a_kind = 0
            hero_villain_count_straight_flush = 0

            hero_ahead_air = 0
            hero_ahead_pair = 0
            hero_ahead_top_pair = 0
            hero_ahead_two_pair = 0
            hero_ahead_overpair = 0
            hero_ahead_trips = 0
            hero_ahead_straight = 0
            hero_ahead_flush = 0
            hero_ahead_full_house = 0
            hero_ahead_four_of_a_kind = 0
            hero_ahead_straight_flush = 0

            #  river fold equity
            river_fold_equity_count = 0

            # 4.  Do testing - get hand count (number of times the hand will be played)

            hand_count = 10

            for i in range(0, hand_count):

                deck = Deck()

                # 4a. Check no overlap of cards between flop, hero, villain, and the deck

                # remove hero cards from deck
                deck.cards.remove(hero_card_1)
                deck.cards.remove(hero_card_2)

                # shuffle villain range combos and pick first

                villain_ranks = [Card.get_rank_int(villain_hand[0]), Card.get_rank_int(villain_hand[1])]

                # remove villain and flop cards from deck (FIXME - why would the cards not be in the deck??)
                deck.cards.remove(villain_hand[0])
                deck.cards.remove(villain_hand[1])
                deck.cards.remove(flop_card_1)
                deck.cards.remove(flop_card_2)
                deck.cards.remove(flop_card_3)
                deck.cards.remove(turn_card[0])


                # ###  Turn

                hero_turn_rank = evaluator.evaluate(hero_hand, turn_board)

                villain_turn_rank = evaluator.evaluate(villain_hand, turn_board)

                if hero_turn_rank < villain_turn_rank:
                    hero_best_turn += 1
                    # if hero_turn_rank <= set_deuces_rating:
                    #     hero_turn_trips_plus = +1
                    #     hero_turn_trips_plus_win = +1

                elif hero_turn_rank == villain_turn_rank:
                    turn_tie += 1
                    # if hero_turn_rank <= set_deuces_rating:
                    #     hero_turn_trips_plus = +1
                    #     hero_turn_trips_plus_tie = +1

                ### hand percentages - start

                if hero_turn_rank < 6186:
                    if hero_turn_rank > 3325:
                        max_board_rank = max(turn_board)
                        max_board_card_rank = Card.get_rank_int(max_board_rank)
                        if max_board_card_rank in hero_ranks:
                            hero_turn_top_pair += 1
                        elif hero_ranks[0] == hero_ranks[1]:
                            if hero_ranks[0] > max_board_card_rank:
                                hero_turn_overpair += 1
                            else:
                                hero_turn_pair += 1
                        else:
                            hero_turn_pair += 1
                    elif hero_turn_rank > 2467:
                        hero_turn_two_pair += 1
                    elif hero_turn_rank > 1609:
                        hero_turn_trips += 1
                    elif hero_turn_rank > 1599:
                        hero_turn_straight += 1
                    elif hero_turn_rank > 322:
                        hero_turn_flush += 1
                    elif hero_turn_rank > 166:
                        hero_turn_full_house += 1
                    elif hero_turn_rank > 10:
                        hero_turn_four_of_a_kind += 1
                    else:
                        hero_turn_straight_flush += 1
                else:  # hero turns air
                    hero_turn_air += 1

                # ### Villain turn analysis

                if villain_turn_rank < 6186:
                    if villain_turn_rank > 3325:
                        max_board_rank = max(turn_board)
                        max_board_card_rank = Card.get_rank_int(max_board_rank)
                        if max_board_card_rank in villain_ranks:
                            villain_turn_top_pair += 1
                        elif hero_ranks[0] == villain_ranks[1]:
                            if hero_ranks[0] > max_board_card_rank:
                                villain_turn_overpair += 1
                            else:
                                villain_turn_pair += 1
                        else:
                            villain_turn_pair += 1
                    elif villain_turn_rank > 2467:
                        villain_turn_two_pair += 1
                    elif villain_turn_rank > 1609:
                        villain_turn_trips += 1
                    elif villain_turn_rank > 1599:
                        villain_turn_straight += 1
                    elif villain_turn_rank > 322:
                        villain_turn_flush += 1
                    elif villain_turn_rank > 166:
                        villain_turn_full_house += 1
                    elif villain_turn_rank > 10:
                        villain_turn_four_of_a_kind += 1
                    else:
                        villain_turn_straight_flush += 1
                else:  # hero turns air
                    villain_turn_air += 1

                villain_turn_air = 0
                villain_turn_full_house = 0
                villain_turn_top_pair = 0
                villain_turn_overpair = 0
                villain_turn_pair = 0
                villain_turn_two_pair = 0
                villain_turn_trips = 0
                villain_turn_straight = 0
                villain_turn_flush = 0
                villain_turn_four_of_a_kind = 0
                villain_turn_straight_flush = 0

                # ###  River

                river_board = turn_board + [deck.draw(1)]

                hero_river_rank = evaluator.evaluate(hero_hand, river_board)
                villain_river_rank = evaluator.evaluate(villain_hand, river_board)

                max_board_rank = max(river_board)
                max_board_card_rank = Card.get_rank_int(max_board_rank)

                hero_ahead_check_count = 0

                if hero_river_rank < villain_river_rank:
                    hero_best_river += 1
                    hero_ahead_check_count = 1
                    # if hero_river_rank <= set_deuces_rating:
                    #     hero_river_trips_plus = +1
                    #     hero_river_trips_plus_win = +1

                elif hero_river_rank == villain_river_rank:
                    river_tie += 1
                    hero_ahead_check_count = 0.5
                    if hero_river_rank <= set_deuces_rating:
                        hero_river_trips_plus = +1
                        hero_river_trips_plus_tie = +1

                if hero_river_rank < 6186:
                    if hero_river_rank > 3325:
                        max_board_rank = max(river_board)
                        max_board_card_rank = Card.get_rank_int(max_board_rank)
                        if max_board_card_rank in hero_ranks:
                            hero_river_top_pair += 1
                            hero_ahead_top_pair += hero_ahead_check_count
                        elif hero_ranks[0] == hero_ranks[1]:
                            if hero_ranks[0] > max_board_card_rank:
                                hero_river_overpair += 1
                                hero_ahead_overpair += hero_ahead_check_count
                            else:
                                hero_river_pair += 1
                                hero_ahead_pair += hero_ahead_check_count
                        else:
                            hero_river_pair += 1
                            hero_ahead_pair += hero_ahead_check_count
                    elif hero_river_rank > 2467:
                        hero_river_two_pair += 1
                        hero_ahead_two_pair += hero_ahead_check_count
                    elif hero_river_rank > 1609:
                        hero_river_trips += 1
                        hero_ahead_trips += hero_ahead_check_count
                    elif hero_river_rank > 1599:
                        hero_river_straight += 1
                        hero_ahead_straight += hero_ahead_check_count
                    elif hero_river_rank > 322:
                        hero_river_flush += 1
                        hero_ahead_flush += hero_ahead_check_count
                    elif hero_river_rank > 166:
                        hero_river_full_house += 1
                        hero_ahead_full_house += hero_ahead_check_count
                    elif hero_river_rank > 10:
                        hero_river_four_of_a_kind += 1
                        hero_ahead_four_of_a_kind += hero_ahead_check_count
                    else:
                        hero_river_straight_flush += 1
                        hero_ahead_straight_flush += hero_ahead_check_count
                else:  # hero river air
                    hero_river_air += 1
                    hero_ahead_air += hero_ahead_check_count

                fold_equity_ranking = evaluator.check_fold_equity(river_board)
                if fold_equity_ranking < villain_river_rank:
                    river_fold_equity_count += 1

            ws1.cell(column=2, row=output_counter).value = hero_card_1
            ws1.cell(column=3, row=output_counter).value = hero_card_2
            ws1.cell(column=4, row=output_counter).value = flop_card_1
            ws1.cell(column=5, row=output_counter).value = flop_card_2
            ws1.cell(column=6, row=output_counter).value = flop_card_3
            ws1.cell(column=7, row=output_counter).value = turn_card[0]
            ws1.cell(column=8, row=output_counter).value = hand_count

            ws1.cell(column=9, row=output_counter).value = turn_tie / hand_count
            ws1.cell(column=10, row=output_counter).value = (hero_best_turn + 0.5 * turn_tie) / hand_count

            ws1.cell(column=11, row=output_counter).value = river_tie / hand_count
            ws1.cell(column=12, row=output_counter).value = (hero_best_river + 0.5 * river_tie) / hand_count

            ws1.cell(column=13, row=output_counter).value = hero_turn_pair / hand_count
            ws1.cell(column=14, row=output_counter).value = hero_turn_top_pair / hand_count
            ws1.cell(column=15, row=output_counter).value = hero_turn_overpair / hand_count
            ws1.cell(column=16, row=output_counter).value = hero_turn_two_pair / hand_count
            ws1.cell(column=17, row=output_counter).value = hero_turn_trips / hand_count
            ws1.cell(column=18, row=output_counter).value = hero_turn_straight / hand_count
            ws1.cell(column=19, row=output_counter).value = hero_turn_flush / hand_count
            ws1.cell(column=20, row=output_counter).value = hero_turn_full_house / hand_count
            ws1.cell(column=21, row=output_counter).value = hero_turn_four_of_a_kind / hand_count
            ws1.cell(column=22, row=output_counter).value = hero_turn_straight_flush / hand_count

            ws1.cell(column=23, row=output_counter).value = hero_river_pair / hand_count
            ws1.cell(column=24, row=output_counter).value = hero_river_top_pair / hand_count
            ws1.cell(column=25, row=output_counter).value = hero_river_overpair / hand_count
            ws1.cell(column=26, row=output_counter).value = hero_river_two_pair / hand_count
            ws1.cell(column=27, row=output_counter).value = hero_river_trips / hand_count
            ws1.cell(column=28, row=output_counter).value = hero_river_straight / hand_count
            ws1.cell(column=29, row=output_counter).value = hero_river_flush / hand_count
            ws1.cell(column=30, row=output_counter).value = hero_river_full_house / hand_count
            ws1.cell(column=31, row=output_counter).value = hero_river_four_of_a_kind / hand_count
            ws1.cell(column=32, row=output_counter).value = hero_river_straight_flush / hand_count

            if hero_river_pair != 0:
                ws1.cell(column=33, row=output_counter).value = hero_ahead_pair / hand_count
            if hero_river_top_pair != 0:
                ws1.cell(column=34, row=output_counter).value = hero_ahead_top_pair / hand_count
            if hero_river_overpair != 0:
                ws1.cell(column=35, row=output_counter).value = hero_ahead_overpair / hand_count
            if hero_river_two_pair != 0:
                ws1.cell(column=36, row=output_counter).value = hero_ahead_two_pair / hand_count
            if hero_river_trips != 0:
                ws1.cell(column=37, row=output_counter).value = hero_ahead_trips / hand_count
            if hero_river_straight != 0:
                ws1.cell(column=38, row=output_counter).value = hero_ahead_straight / hand_count
            if hero_river_flush != 0:
                ws1.cell(column=39, row=output_counter).value = hero_ahead_flush / hand_count
            if hero_river_full_house != 0:
                ws1.cell(column=40, row=output_counter).value = hero_ahead_full_house / hand_count
            if hero_river_four_of_a_kind != 0:
                ws1.cell(column=41, row=output_counter).value = hero_ahead_four_of_a_kind / hand_count
            if hero_river_straight_flush != 0:
                ws1.cell(column=42, row=output_counter).value = hero_ahead_straight_flush / hand_count

            ws1.cell(column=43, row=output_counter).value = hero_turn_air / hand_count
            ws1.cell(column=44, row=output_counter).value = hero_river_air / hand_count
            ws1.cell(column=45, row=output_counter).value = hero_ahead_air / hand_count

            ##  new end

            ws1.cell(column=46, row=output_counter).value = river_fold_equity_count / hand_count

            ws1.cell(column=47, row=output_counter).value = villain_combo
            ws1.cell(column=48, row=output_counter).value = villain_actual
            # ws1.cell(column=49, row=output_counter).value = villain_card_1
            # ws1.cell(column=50, row=output_counter).value = villain_card_2

            output_counter += 1
            print(output_counter)
wb_output.save(filename=dest_filename)