#  Purpose of this module is to test against an opening range - One Raiser in Front

import random
from evaluator import Evaluator
from deck import Deck
from openpyxl import load_workbook
from card import Card
import e_functions as matt_custom

# 1. Initiate class

evaluator = Evaluator()

# 2. Workbooks used for outpu
# t and input

# 2a Input workbook is range_lookup.xlsm
wb_input = load_workbook(filename='combo_method.xlsm', data_only=True)
input_sheet = wb_input['preflop']

# 2b Output workbook
dest_filename = 'combo_method_output.xlsx'
wb_output = load_workbook(filename=dest_filename, read_only=False)
ws1 = wb_output['preflop']

ws_all_combos = wb_input['all_combos']

output_counter = ws1.max_row + 1

hero_combo_count = input_sheet.cell(row=2, column=3).value

for l in range(5, hero_combo_count+5):

    # number of hero cards to test
    set_deuces_rating = 2467  # ranking of three deuces
    hero_combo = input_sheet.cell(row=l, column=2).value
    hero_actual = input_sheet.cell(row=l, column=3).value
    hero_card_1 = input_sheet.cell(row=l, column=4).value
    hero_card_2 = input_sheet.cell(row=l, column=5).value
    hero_hand = [hero_card_1, hero_card_2]
    hero_ranks = [Card.get_rank_int(hero_card_1), Card.get_rank_int(hero_card_2)]

    for j in range(2, 1327):

        hand_count = 0
        hero_best_flop = 0
        hero_best_turn = 0
        hero_best_river = 0
        flop_tie = 0
        turn_tie = 0
        river_tie = 0

        # drawing power
        # flop
        hero_flop_air = 0
        hero_flop_ace_high = 0
        hero_flop_pair = 0
        hero_flop_top_pair = 0
        hero_flop_overpair = 0
        hero_flop_two_pair = 0
        hero_flop_trips = 0
        hero_flop_straight = 0
        hero_flop_flush = 0
        hero_flop_full_house = 0
        hero_flop_four_of_a_kind = 0
        hero_flop_straight_flush = 0

        # turn
        hero_turn_air = 0
        hero_turn_ace_high = 0
        hero_turn_pair = 0
        hero_turn_top_pair = 0
        hero_turn_overpair = 0
        hero_turn_two_pair = 0
        hero_turn_trips = 0
        hero_turn_straight = 0
        hero_turn_flush = 0
        hero_turn_full_house = 0
        hero_turn_four_of_a_kind = 0
        hero_turn_straight_flush = 0

        # river
        hero_river_air = 0
        hero_river_ace_high = 0
        hero_river_pair = 0
        hero_river_top_pair = 0
        hero_river_overpair = 0
        hero_river_two_pair = 0
        hero_river_trips = 0
        hero_river_straight = 0
        hero_river_flush = 0
        hero_river_full_house = 0
        hero_river_four_of_a_kind = 0
        hero_river_straight_flush = 0

        # hand strength

        hero_villain_count_pair = 0
        hero_villain_count_top_pair = 0
        hero_villain_count_two_pair = 0
        hero_villain_count_overpair = 0
        hero_villain_count_trips = 0
        hero_villain_count_straight = 0
        hero_villain_count_flush = 0
        hero_villain_count_full_house = 0
        hero_villain_count_four_of_a_kind = 0
        hero_villain_count_straight_flush = 0

        flop_hero_ahead_air = 0
        flop_hero_ahead_ace_high = 0
        flop_hero_ahead_pair = 0
        flop_hero_ahead_top_pair = 0
        flop_hero_ahead_two_pair = 0
        flop_hero_ahead_overpair = 0
        flop_hero_ahead_trips = 0
        flop_hero_ahead_straight = 0
        flop_hero_ahead_flush = 0
        flop_hero_ahead_full_house = 0
        flop_hero_ahead_four_of_a_kind = 0
        flop_hero_ahead_straight_flush = 0

        turn_hero_ahead_air = 0
        turn_hero_ahead_ace_high = 0
        turn_hero_ahead_pair = 0
        turn_hero_ahead_top_pair = 0
        turn_hero_ahead_two_pair = 0
        turn_hero_ahead_overpair = 0
        turn_hero_ahead_trips = 0
        turn_hero_ahead_straight = 0
        turn_hero_ahead_flush = 0
        turn_hero_ahead_full_house = 0
        turn_hero_ahead_four_of_a_kind = 0
        turn_hero_ahead_straight_flush = 0

        hero_ahead_air = 0
        hero_ahead_ace_high = 0
        hero_ahead_pair = 0
        hero_ahead_top_pair = 0
        hero_ahead_two_pair = 0
        hero_ahead_overpair = 0
        hero_ahead_trips = 0
        hero_ahead_straight = 0
        hero_ahead_flush = 0
        hero_ahead_full_house = 0
        hero_ahead_four_of_a_kind = 0
        hero_ahead_straight_flush = 0

        villain_combo = ws_all_combos.cell(row=j, column=2).value
        villain_actual = ws_all_combos.cell(row=j, column=3).value
        villain_card_1 = ws_all_combos.cell(row=j, column=6).value
        villain_card_2 = ws_all_combos.cell(row=j, column=7).value
        villain_hand = [villain_card_1, villain_card_2]

        empty_hero_hand = []

        if hero_card_1 not in villain_hand and hero_card_2 not in villain_hand:

            hand_count = 50

            for i in range(0, hand_count):

                deck = Deck()

                # 4a. Check no overlap of cards between flop, hero, villain, and the deck

                # remove hero cards from deck
                deck.cards.remove(hero_card_1)
                deck.cards.remove(hero_card_2)

                deck.cards.remove(villain_hand[0])
                deck.cards.remove(villain_hand[1])

                all_board_cards = deck.draw(5)

                ######### FLOP #########
                board = all_board_cards[:3]

                hero_flop_rank = evaluator.evaluate(hero_hand, board)
                villain_flop_rank = evaluator.evaluate(villain_hand, board)

                hero_ahead_check_count = 0

                if hero_flop_rank < villain_flop_rank:
                    hero_best_flop += 1
                    hero_ahead_check_count = 1
                elif hero_flop_rank == villain_flop_rank:
                    flop_tie += 1
                    hero_ahead_check_count = 0.5

                if hero_flop_rank < 6186:
                    if hero_flop_rank > 3325:
                        if matt_custom.flop_check_for_board_pair(board):
                            if max(hero_ranks) == 12:
                                hero_flop_ace_high += 1
                                flop_hero_ahead_ace_high += hero_ahead_check_count
                            else:
                                hero_flop_air += 1
                                flop_hero_ahead_air += hero_ahead_check_count
                        else:
                            max_board_rank = max(board)
                            max_board_card_rank = Card.get_rank_int(max_board_rank)
                            if max_board_card_rank in hero_ranks:
                                hero_flop_top_pair += 1
                                flop_hero_ahead_top_pair += hero_ahead_check_count
                            elif hero_ranks[0] == hero_ranks[1]:
                                if hero_ranks[0] > max_board_card_rank:
                                    hero_flop_overpair += 1
                                    flop_hero_ahead_overpair += hero_ahead_check_count
                                else:
                                    hero_flop_pair += 1
                                    flop_hero_ahead_pair += hero_ahead_check_count
                            else:
                                hero_flop_pair += 1
                    elif hero_flop_rank > 2467:
                        if matt_custom.flop_check_for_board_pair(board):
                            max_board_rank = max(board)  # FIXME what if top card is a pair of three
                            max_board_card_rank = Card.get_rank_int(max_board_rank)
                            if max_board_card_rank in hero_ranks:
                                hero_flop_top_pair += 1
                                flop_hero_ahead_top_pair += hero_ahead_check_count
                            elif hero_ranks[0] == hero_ranks[1]:
                                if hero_ranks[0] > max_board_card_rank:
                                    hero_flop_overpair += 1
                                    flop_hero_ahead_overpair += hero_ahead_check_count
                                else:
                                    hero_flop_pair += 1
                                    flop_hero_ahead_pair += hero_ahead_check_count
                            else:
                                hero_flop_pair += 1
                                flop_hero_ahead_pair += hero_ahead_check_count
                        else:
                            hero_flop_two_pair += 1
                            flop_hero_ahead_two_pair += hero_ahead_check_count
                    elif hero_flop_rank > 1609:
                        if matt_custom.flop_check_for_board_trips(board):
                            if max(hero_ranks) == 12:
                                hero_flop_ace_high += 1
                                flop_hero_ahead_air += hero_ahead_check_count
                            else:
                                hero_flop_air += 1
                                flop_hero_ahead_air += hero_ahead_check_count
                        else:
                            hero_flop_trips += 1
                            flop_hero_ahead_trips += hero_ahead_check_count
                    elif hero_flop_rank > 1599:
                        hero_flop_straight += 1
                        flop_hero_ahead_straight += hero_ahead_check_count
                    elif hero_flop_rank > 322:
                        hero_flop_flush += 1
                        flop_hero_ahead_flush += hero_ahead_check_count
                    elif hero_flop_rank > 166:
                        hero_flop_full_house += 1
                        flop_hero_ahead_full_house += hero_ahead_check_count
                    elif hero_flop_rank > 10:
                        hero_flop_four_of_a_kind += 1
                        flop_hero_ahead_four_of_a_kind += hero_ahead_check_count
                    else:
                        hero_flop_straight_flush += 1
                        flop_hero_ahead_straight_flush += hero_ahead_check_count
                else:  # hero flops air
                    if max(hero_ranks) == 12:
                        hero_flop_ace_high += 1
                        flop_hero_ahead_air += hero_ahead_check_count
                    else:
                        hero_flop_air += 1
                        flop_hero_ahead_air += hero_ahead_check_count

                ######### TURN #########

                board = all_board_cards[:4]

                hero_turn_rank = evaluator.evaluate(hero_hand, board)
                villain_turn_rank = evaluator.evaluate(villain_hand, board)

                if hero_turn_rank < villain_turn_rank:
                    hero_best_turn += 1
                elif hero_turn_rank == villain_turn_rank:
                    turn_tie += 1

                if hero_turn_rank < 6186:
                    if hero_turn_rank > 3325:
                        if matt_custom.turn_check_for_board_pair(board):
                            if max(hero_ranks) == 12:
                                hero_turn_ace_high += 1
                            else:
                                hero_turn_air += 1

                        else:
                            max_board_rank = max(board)  # FIXME what if top card is a pair of three
                            max_board_card_rank = Card.get_rank_int(max_board_rank)
                            if max_board_card_rank in hero_ranks:
                                hero_turn_top_pair += 1
                            elif hero_ranks[0] == hero_ranks[1]:
                                if hero_ranks[0] > max_board_card_rank:
                                    hero_turn_overpair += 1
                                else:
                                    hero_turn_pair += 1
                            else:
                                hero_turn_pair += 1
                    elif hero_turn_rank > 2467:
                        two_pair_check = matt_custom.turn_check_for_board_two_pairs(hero_hand, board)
                        if two_pair_check == "two_pair":
                            hero_turn_two_pair += 1
                        elif two_pair_check == "overpair":
                            hero_turn_overpair += 1
                        elif two_pair_check == "top_pair":
                            hero_turn_top_pair += 1
                        elif two_pair_check == "pair":
                            hero_turn_pair += 1
                        elif two_pair_check == "air":
                            if max(hero_ranks) == 12:
                                hero_turn_ace_high += 1
                            else:
                                hero_turn_air += 1
                    elif hero_turn_rank > 1609:
                        if matt_custom.turn_check_for_board_trips(board):
                            if max(hero_ranks) == 12:
                                hero_turn_ace_high += 1
                            else:
                                hero_turn_air += 1
                        else:
                            hero_turn_trips += 1
                    elif hero_turn_rank > 1599:
                        hero_turn_straight += 1
                    elif hero_turn_rank > 322:
                        hero_turn_flush += 1
                    elif hero_turn_rank > 166:
                        hero_turn_full_house += 1
                    elif hero_turn_rank > 10:
                        hero_turn_four_of_a_kind += 1
                    else:
                        hero_turn_straight_flush += 1
                else:  # hero turns air
                    if max(hero_ranks) == 12:
                        hero_turn_ace_high += 1
                    else:
                        hero_turn_air += 1

                ######### RIVER #########
                # evaluator.eliminate_board_effect(hand_type, board, hand)
                board = all_board_cards[:5]

                hero_river_rank = evaluator.evaluate(hero_hand, board)
                villain_river_rank = evaluator.evaluate(villain_hand, board)

                hero_ahead_check_count = 0

                if hero_river_rank < villain_river_rank:
                    hero_best_river += 1
                    hero_ahead_check_count = 1
                elif hero_river_rank == villain_river_rank:
                    river_tie += 1
                    hero_ahead_check_count = 0.5
                # work out how many beat you!

                if hero_river_rank < 6186:
                    if hero_river_rank > 3325:
                        if matt_custom.river_check_for_board_pair(board):
                            if max(hero_ranks) == 12:
                                hero_river_ace_high += 1
                                hero_ahead_ace_high += hero_ahead_check_count
                            else:
                                hero_river_air += 1
                                hero_ahead_air += hero_ahead_check_count
                        else:
                            max_board_rank = max(board)  # FIXME what if top card is a pair of three
                            max_board_card_rank = Card.get_rank_int(max_board_rank)
                            if max_board_card_rank in hero_ranks:
                                hero_river_top_pair += 1
                                hero_ahead_top_pair += hero_ahead_check_count
                            elif hero_ranks[0] == hero_ranks[1]:
                                if hero_ranks[0] > max_board_card_rank:
                                    hero_river_overpair += 1
                                    hero_ahead_overpair += hero_ahead_check_count
                                else:
                                    hero_river_pair += 1
                                    hero_ahead_pair += hero_ahead_check_count
                            else:
                                hero_river_pair += 1
                                hero_ahead_pair += hero_ahead_check_count

                    elif hero_river_rank > 2467:
                        two_pair_check = matt_custom.turn_check_for_board_two_pairs(hero_hand, board)
                        if two_pair_check == "two_pair":
                            hero_river_two_pair += 1
                            hero_ahead_two_pair += hero_ahead_check_count
                        elif two_pair_check == "overpair":
                            hero_river_overpair += 1
                            hero_ahead_overpair += hero_ahead_check_count
                        elif two_pair_check == "top_pair":
                            hero_river_top_pair += 1
                            hero_ahead_top_pair += hero_ahead_check_count
                        elif two_pair_check == "pair":
                            hero_river_pair += 1
                            hero_ahead_pair += hero_ahead_check_count
                        elif two_pair_check == "air":
                            if max(hero_ranks) == 12:
                                hero_river_ace_high += 1
                                hero_ahead_ace_high += hero_ahead_check_count
                            else:
                                hero_river_air += 1
                                hero_ahead_air += hero_ahead_check_count

                    elif hero_river_rank > 1609:
                        if matt_custom.river_check_for_board_trips(board):
                            if max(hero_ranks) == 12:
                                hero_river_ace_high += 1
                                hero_ahead_ace_high += hero_ahead_check_count
                            else:
                                hero_river_air += 1
                                hero_ahead_air += hero_ahead_check_count
                        else:
                            hero_river_trips += 1
                            hero_ahead_trips += hero_ahead_check_count

                    elif hero_river_rank > 1599:
                        if evaluator.evaluate(empty_hero_hand, board) > hero_river_rank:
                            hero_river_air += 1
                            hero_ahead_air += hero_ahead_check_count
                        else:
                            hero_river_straight += 1
                            hero_ahead_straight += hero_ahead_check_count
                    elif hero_river_rank > 322:
                        if evaluator.evaluate(empty_hero_hand, board) > hero_river_rank:
                            hero_river_air += 1
                            hero_ahead_air += hero_ahead_check_count
                        else:
                            hero_river_flush += 1
                            hero_ahead_flush += hero_ahead_check_count
                    elif hero_river_rank > 166:
                        if evaluator.evaluate(empty_hero_hand, board) > hero_river_rank:
                            hero_river_air += 1
                            hero_ahead_air += hero_ahead_check_count
                        else:
                            hero_river_full_house += 1
                            hero_ahead_full_house += hero_ahead_check_count
                    elif hero_river_rank > 10:
                        if evaluator.evaluate(empty_hero_hand, board) < 166:
                            hero_river_air += 1
                            hero_ahead_air += hero_ahead_check_count
                        else:
                            hero_river_four_of_a_kind += 1
                            hero_ahead_four_of_a_kind += hero_ahead_check_count
                    else:
                        if evaluator.evaluate(empty_hero_hand, board) > hero_river_rank:
                            hero_river_air += 1
                            hero_ahead_air += hero_ahead_check_count
                        else:
                            hero_river_straight_flush += 1
                            hero_ahead_straight_flush += hero_ahead_check_count
                else:  # hero river air
                    if max(hero_ranks) == 12:
                        hero_river_ace_high += 1
                        hero_ahead_ace_high += hero_ahead_check_count
                    else:
                        hero_river_air += 1
                        hero_ahead_air += hero_ahead_check_count

            ws1.cell(column=1, row=output_counter).value = hero_actual
            ws1.cell(column=2, row=output_counter).value = villain_combo

            # ws1.cell(column=2, row=output_counter).value = hero_card_1
            ws1.cell(column=3, row=output_counter).value = villain_actual
            ws1.cell(column=4, row=output_counter).value = hand_count
            ws1.cell(column=5, row=output_counter).value = (hero_best_flop + 0.5 * flop_tie) / hand_count
            ws1.cell(column=6, row=output_counter).value = (hero_best_turn + 0.5 * turn_tie) / hand_count
            ws1.cell(column=7, row=output_counter).value = (hero_best_river + 0.5 * river_tie) / hand_count

            ws1.cell(column=8, row=output_counter).value = hero_flop_pair / hand_count
            ws1.cell(column=9, row=output_counter).value = hero_flop_top_pair / hand_count
            ws1.cell(column=10, row=output_counter).value = hero_flop_overpair / hand_count
            ws1.cell(column=11, row=output_counter).value = hero_flop_two_pair / hand_count
            ws1.cell(column=12, row=output_counter).value = hero_flop_trips / hand_count
            ws1.cell(column=13, row=output_counter).value = hero_flop_straight / hand_count
            ws1.cell(column=14, row=output_counter).value = hero_flop_flush / hand_count
            ws1.cell(column=15, row=output_counter).value = hero_flop_full_house / hand_count
            ws1.cell(column=16, row=output_counter).value = hero_flop_four_of_a_kind / hand_count
            ws1.cell(column=17, row=output_counter).value = hero_flop_straight_flush / hand_count

            ws1.cell(column=18, row=output_counter).value = hero_turn_pair / hand_count
            ws1.cell(column=19, row=output_counter).value = hero_turn_top_pair / hand_count
            ws1.cell(column=20, row=output_counter).value = hero_turn_overpair / hand_count
            ws1.cell(column=21, row=output_counter).value = hero_turn_two_pair / hand_count
            ws1.cell(column=22, row=output_counter).value = hero_turn_trips / hand_count
            ws1.cell(column=23, row=output_counter).value = hero_turn_straight / hand_count
            ws1.cell(column=24, row=output_counter).value = hero_turn_flush / hand_count
            ws1.cell(column=25, row=output_counter).value = hero_turn_full_house / hand_count
            ws1.cell(column=26, row=output_counter).value = hero_turn_four_of_a_kind / hand_count
            ws1.cell(column=27, row=output_counter).value = hero_turn_straight_flush / hand_count

            ws1.cell(column=28, row=output_counter).value = hero_river_pair / hand_count
            ws1.cell(column=29, row=output_counter).value = hero_river_top_pair / hand_count
            ws1.cell(column=30, row=output_counter).value = hero_river_overpair / hand_count
            ws1.cell(column=31, row=output_counter).value = hero_river_two_pair / hand_count
            ws1.cell(column=32, row=output_counter).value = hero_river_trips / hand_count
            ws1.cell(column=33, row=output_counter).value = hero_river_straight / hand_count
            ws1.cell(column=34, row=output_counter).value = hero_river_flush / hand_count
            ws1.cell(column=35, row=output_counter).value = hero_river_full_house / hand_count
            ws1.cell(column=36, row=output_counter).value = hero_river_four_of_a_kind / hand_count
            ws1.cell(column=37, row=output_counter).value = hero_river_straight_flush / hand_count

            if hero_river_pair != 0:
                ws1.cell(column=38, row=output_counter).value = hero_ahead_pair / hand_count
            if hero_river_top_pair != 0:
                ws1.cell(column=39, row=output_counter).value = hero_ahead_top_pair / hand_count
            if hero_river_overpair != 0:
                ws1.cell(column=40, row=output_counter).value = hero_ahead_overpair / hand_count
            if hero_river_two_pair != 0:
                ws1.cell(column=41, row=output_counter).value = hero_ahead_two_pair / hand_count
            if hero_river_trips != 0:
                ws1.cell(column=42, row=output_counter).value = hero_ahead_trips / hand_count
            if hero_river_straight != 0:
                ws1.cell(column=43, row=output_counter).value = hero_ahead_straight / hand_count
            if hero_river_flush != 0:
                ws1.cell(column=44, row=output_counter).value = hero_ahead_flush / hand_count
            if hero_river_full_house != 0:
                ws1.cell(column=45, row=output_counter).value = hero_ahead_full_house / hand_count
            if hero_river_four_of_a_kind != 0:
                ws1.cell(column=46, row=output_counter).value = hero_ahead_four_of_a_kind / hand_count
            if hero_river_straight_flush != 0:
                ws1.cell(column=47, row=output_counter).value = hero_ahead_straight_flush / hand_count

            # ws1.cell(column=48, row=output_counter).value = flop_count

            ws1.cell(column=48, row=output_counter).value = hero_flop_air / hand_count
            ws1.cell(column=49, row=output_counter).value = hero_turn_air / hand_count
            ws1.cell(column=50, row=output_counter).value = hero_river_air / hand_count
            ws1.cell(column=51, row=output_counter).value = hero_ahead_air / hand_count

            ws1.cell(column=52, row=output_counter).value = flop_tie / hand_count
            ws1.cell(column=53, row=output_counter).value = turn_tie / hand_count
            ws1.cell(column=54, row=output_counter).value = river_tie / hand_count

            ws1.cell(column=55, row=output_counter).value = hero_combo

            output_counter += 1
            print(output_counter)

    wb_output.save(filename=dest_filename)


